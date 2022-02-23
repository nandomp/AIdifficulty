

import os
import pandas as pd

import openpyxl
files=[
    "/home/dcast/adversarial_project/irt_to_nlp/AAAI_Results/WMProb/IMDb/PLOTS.xlsx",
    "/home/dcast/adversarial_project/irt_to_nlp/AAAI_Results/WMProb/SST-2/PLOTS.xlsx",
    "/home/dcast/adversarial_project/irt_to_nlp/AAAI_Results/WSBias/SST-2/PLOTS-872.xlsx"
       ]

def get_df_from_xlsx_path(xlsx_path):
    workbook=openpyxl.load_workbook(xlsx_path)

    sheetnames=workbook.sheetnames
    label_by_sheet={}
    for sheet in sheetnames:
        data=pd.read_excel(xlsx_path,sheet_name=sheet)
        try:
            label_by_sheet[sheet]=data.LABEL
            print(f"the model {sheet} work")
        except:
            print("this sheet don't work")
            
    # print(label_by_sheet)

    data=pd.DataFrame(label_by_sheet)

    # print(data.head())
    return data


for fpath in files:
    
    splitter_fpath=fpath.split("/")
    fname=splitter_fpath[-1].split(".")[0]
    dataset=splitter_fpath[-2]
    main_folder=splitter_fpath[-3]
    new_fpath=os.path.join("/home/dcast/adversarial_project/irt_to_nlp/AAAI_Results/results_to_nando",
                           main_folder+"_"+dataset+"_"+fname+".csv")
    data=get_df_from_xlsx_path(fpath)
    
    
    data.to_csv(new_fpath)